from faker import Faker
import pandas as pd
from alive_progress import alive_bar
from office.lib.utils import pandas_mem
import os
from pathlib import Path
from openpyxl import load_workbook
import warnings

# 忽略waring警告
warnings.filterwarnings("ignore")


class MainExcel():

    def fake2excel(self, columns, rows, path, language):
        """
        @Author & Date  : CoderWanFeng 2022/5/13 0:12
        @Desc  : columns:list，每列的数据名称，默认是名称
                rows：多少行，默认是1
                language：什么语言，可以填english，默认是中文
                path：输出excel的位置，有默认值
        """
        # 可以选择英语
        if language.lower() == 'english':
            language = 'en_US'
        else:
            language = 'zh_CN'
        # 开始造数
        fake = Faker(language)
        excel_dict = {}
        with alive_bar(len(columns) * rows) as bar:
            for column in columns:
                excel_dict[column] = list()
                # excel_dict[column] = map(lambda x: eval('fake.{func}()'.format(func=x)), [column] * rows) # 使用map，会报错
                while len(excel_dict[column]) < rows:
                    excel_dict[column].append(eval('fake.{func}()'.format(func=column)))
                    bar()
            # 用pandas，将模拟数据，写进excel里面
            writer = pd.ExcelWriter(path)
            data = pd.DataFrame(excel_dict)
            data = pandas_mem.reduce_pandas_mem_usage(data)
            data.to_excel(writer, index=False)
            # writer.save()
            writer.close()

    def merge2excel(self, dir_path, output_file, xlsxSuffix=".xlsx"):
        """
        :param dir_path:
        :param output_file:
        :return:
        """
        if not output_file.endswith(xlsxSuffix):
            raise Exception(f'您自定义的输出文件名，不是以{xlsxSuffix}结尾的')
        file_path_dict = self.getfile(dir_path)  # excel文件所在的文件夹
        try:
            writer = pd.ExcelWriter(output_file)  # 合并后的excel名称
        except PermissionError:
            raise Exception(f'小可爱，你的输出文件，是不是上次打开了没关闭呀？这是你自己指定的输出文件名称：{output_file}')
        for file, path in file_path_dict.items():
            if file.endswith("xlsx"):
                df = pd.read_excel(path)
            if file.endswith("csv"):
                df = pd.read_csv(path)
            df.to_excel(writer, sheet_name=file.split('.')[0], index=False)
        print(f'您指定的Excel文件已经合并完毕，合并后的文件名是{output_file}')
        writer.save()

    def getfile(self, dirpath):
        path = Path(dirpath)
        file_path_dict = {}
        for root, dirs, files in os.walk(dirpath):
            for file in files:
                if file.endswith("xlsx") or file.endswith("csv"):
                    file_path_dict[file] = (path / file)
        return file_path_dict

    def sheet2excel(self, file_path):
        # 先读取一次文件，获取sheet表的名称

        origin_excel = load_workbook(filename=file_path)  # 读取原excel文件
        origin_sheet_names = origin_excel.sheetnames  # 获取sheet的名称
        print(f'一共有{len(origin_sheet_names)}个sheet，名称分别为：{origin_sheet_names}')
        print('拆分开始')

        if len(origin_sheet_names) > 1:  # 如果sheetnames小于1，报错：该文件不需要拆分

            for j in range(len(origin_sheet_names)):

                wb = load_workbook(filename=file_path)  # 再读取一次文件，由于每次删除后需要保存一次，所以不能与上一次一样
                sheet = wb[origin_sheet_names[j]]
                wb.copy_worksheet(sheet)

                new_filename = origin_sheet_names[j] + '.xlsx'  # 新建一个sheet命名的excel文件

                for i in range(len(origin_sheet_names)):
                    sheet1 = wb[origin_sheet_names[i]]
                    wb.remove(sheet1)

                wb.save(filename=new_filename)

                # 由于使用copy_worksheet后，sheet表名有copy字段，这里做个调整

                new = load_workbook(filename=new_filename)
                news = new.active
                news.title = origin_sheet_names[j]
                new.save(filename=new_filename)
            print('拆分结束')
        else:
            raise Exception(f"你的文件只有一个sheet，难道还要拆分吗？我做不到啊~~~，你的文件名{file_path}")

    def merge2sheet(self, dir_path, output_sheet_name: str, output_excel_name):
        for root, dirs, files in os.walk(dir_path):
            path = Path(dir_path)
            print(f'正在合并的文件有：{files}')
            print(f'合并后的文件名是：{output_excel_name}')
            print(f'合并后的sheet名是：{output_sheet_name}')
            df_list = []
            for file in files:
                if file.endswith("xlsx") or file.endswith("xls"):
                    excel_path = (path / file)
                    df_list.append(pd.read_excel(excel_path))
            res = pd.concat(df_list)
            res.to_excel(
                (path / (output_excel_name + '.xlsx')),
                sheet_name=output_sheet_name,
                index=False  # 不保留index
            )

            pass
