import os
import xlrd, xlwt
import openpyxl
import datetime


#

def generate_xls(filepath: str, worksheet_data: dict) -> str:
    """生成XLS格式的拆分文件。
    
    Args:
        filepath (str): 原始文件路径
        worksheet_data (dict): 工作表数据字典
    
    Returns:
        str: 新生成的文件路径
    """
    datetime_str = datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')
    new_filepath = filepath.replace('.xls', '_Split_{}.xls'.format(datetime_str))
    new_workbook = xlwt.Workbook(encoding='utf-8')
    for worksheet_name, row_data_list in worksheet_data.items():
        new_worksheet = new_workbook.add_sheet(worksheet_name)
        for row_index, row_data in enumerate(row_data_list):
            for column_index, data in enumerate(row_data):
                new_worksheet.write(row_index, column_index, data)
    new_workbook.save(new_filepath)
    return new_filepath


def process_xls(filepath: str, column: int, worksheet_name: str = None) -> str:
    """处理XLS格式的Excel文件拆分。
    
    Args:
        filepath (str): Excel文件路径
        column (int): 拆分依据的列号
        worksheet_name (str, optional): 工作表名称，默认为None
    
    Returns:
        str: 处理结果信息
    """
    try:
        workbook = xlrd.open_workbook(filepath, formatting_info=True)
    except:
        return "文件读取异常：{}".format(filepath)
    if worksheet_name:
        worksheet = workbook.sheet_by_name(worksheet_name)
    else:
        worksheet = workbook.sheet_by_index(0)
    rows = worksheet.nrows
    cols = worksheet.ncols
    split_data_dict = {}
    for r in tqdm(range(rows)):
        row_data = [worksheet.cell(r, c).value if worksheet.cell(r, c).value else ' ' for c in range(cols)]
        temp_data = row_data[column - 1]
        temp_data_list = split_data_dict.get(temp_data, [])
        temp_data_list.append(row_data)
        split_data_dict[temp_data] = temp_data_list
    new_filepath = generate_xls(filepath, split_data_dict)
    return "数据保存在新文件中，文件名：{}".format(new_filepath)


def generate_xlsx(filepath: str, worksheet_data: dict) -> str:
    """生成XLSX格式的拆分文件。
    
    Args:
        filepath (str): 原始文件路径
        worksheet_data (dict): 工作表数据字典
    
    Returns:
        str: 新生成的文件路径
    """
    datetime_str = datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')
    new_filepath = filepath.replace('.xlsx', '_Split_{}.xlsx'.format(datetime_str))
    new_workbook = openpyxl.Workbook()
    for worksheet_name, row_data_list in worksheet_data.items():
        new_worksheet = new_workbook.create_sheet(worksheet_name)
        for row_data in row_data_list:
            new_worksheet.append(row_data)
    new_workbook.save(new_filepath)
    return new_filepath


def process_xlsx(filepath: str, column: int, worksheet_name: str = None) -> str:
    """处理XLSX格式的Excel文件拆分。
    
    Args:
        filepath (str): Excel文件路径
        column (int): 拆分依据的列号
        worksheet_name (str, optional): 工作表名称，默认为None
    
    Returns:
        str: 处理结果信息
    """
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except:
        return "文件读取异常：{}".format(filepath)
    if worksheet_name:
        worksheet = workbook.get_sheet_by_name(worksheet_name)
    else:
        worksheet = workbook.active
    if worksheet.max_column < column:
        return "最大列数是{}，取不到第{}列".format(worksheet.max_column, column)

    split_data_dict = {}
    for row in worksheet.rows:
        row_data = [cell.value if cell.value else ' ' for cell in row]
        temp_data = row_data[column - 1]
        temp_data_list = split_data_dict.get(temp_data, [])
        temp_data_list.append(row_data)
        split_data_dict[temp_data] = temp_data_list
    new_filepath = generate_xlsx(filepath, split_data_dict)
    return "数据保存在新文件中，文件名：{}".format(new_filepath)


def split_excel_by_column(filepath: str, column: int, worksheet_name: str = None) -> str:
    """根据指定列拆分Excel文件。
    
    Args:
        filepath (str): Excel文件路径
        column (int): 拆分依据的列号
        worksheet_name (str, optional): 工作表名称，默认为None
    
    Returns:
        str: 处理结果信息
    """
    if filepath.endswith('.xlsx'):
        result = process_xlsx(filepath, column, worksheet_name)
    elif filepath.endswith('.xls'):
        result = process_xls(filepath, column, worksheet_name)
    else:
        print("文件格式不对，不进行处理")
        return "文件格式不对，不进行处理"
    print(result)
    return result


if __name__ == "__main__":
    filename = 'sedemo.xls'
    # filename = 'SEdemo.xlsx'
    result = split_excel_by_column(filename, 6)  # 处理文件，表格的第六列，worksheet_name指定工作表，不指定则读取文件默认工作表
    print(result)
