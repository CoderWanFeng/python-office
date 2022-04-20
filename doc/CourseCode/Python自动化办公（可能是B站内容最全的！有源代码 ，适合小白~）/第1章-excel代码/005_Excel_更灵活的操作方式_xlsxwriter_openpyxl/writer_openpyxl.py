# -*- coding: utf-8 -*-
# @Time : 2020/8/20 17:16
# @公众号 :Python自动化办公社区 
# @File : writer_openpyxl.py
# @Software: PyCharm
# @Description:

# import xlwt
#
# workbook = xlwt.Workbook()
# sheet0 = workbook.add_sheet('sheet0')
# for i in range(0,300):
#     sheet0.write(0,i,i)
# workbook.save('num.xls')

# 不带格式
import xlsxwriter as xw
workbook = xw.Workbook('number.xlsx')
sheet0 = workbook.add_worksheet('sheet0')
for i in range(0,300):
    sheet0.write(0,i,i)
workbook.close()


# 性能不稳定
import openpyxl
workbook = openpyxl.load_workbook('number.xlsx')
sheet0 = workbook['sheet0']
sheet0['B3']= '2'
sheet0['C2']= '4'
sheet0['D7']= '3'
workbook.save('num_open.xlsx')

