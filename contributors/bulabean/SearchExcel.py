import os
import openpyxl
import xlrd
import datetime
import time


def change_datatype(row_data: list):
    """
    excel单元格的内容类型检测和转换
    参数：
        row_data：行数据，列表格式
    """
    result_data = []
    for rd in row_data:
        if type(rd) == datetime.datetime:
            t = rd.strftime("%Y-%m-%d %H:%M:%S")
        elif type(rd) == str:
            t = rd
        elif type(rd) == int:
            t = str(rd)
        elif type(rd) == float:
            t = str(rd)
        elif type(rd) is None:
            t = ''
        else:
            t = str(rd)
        result_data.append(t)
    return result_data


def find_key(search_key: str, row_content: str):
    """
    检测关键词和内容
    参数：
        search_key：关键词
        row_content：行内容
    """
    if search_key in row_content:
        return True
    else:
        return False


def process_xls(path, file):
    """
    读取xls后缀的excel文件
    参数：
        path：文件所在路径
        file：文件名
    """
    filepath = os.path.join(path, file)
    try:
        rb = xlrd.open_workbook(filepath, formatting_info=True)
    except:
        return False
    sheet_names = rb.sheet_names()
    space_line = 0
    for ws_name in sheet_names:
        ws = rb.sheet_by_name(ws_name)
        rows = ws.nrows
        cols = ws.ncols
        for r in range(rows):
            values = [ws.cell(r, c).value for c in range(cols)]
            values = change_datatype(values)
            values = " ".join(values)
            if values:
                yield filepath, ws_name, r, values  # 文件路径，工作表名，行数，行内容
            else:
                if space_line < 10:
                    space_line += 1
                else:
                    break


def process_xlsx(path, file):
    """
    读取xlsx后缀的excel文件
    参数：
        path：文件所在路径
        file：文件名
    """
    filepath = os.path.join(path, file)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except:
        return False
    worksheets_name = wb.sheetnames
    space_line = 0
    for ws_name in worksheets_name:
        ws = wb[ws_name]
        for index, row in enumerate(ws.rows):
            values = [r.value for r in row if r.value != None]
            values = change_datatype(values)
            values = " ".join(values)
            if values:
                yield filepath, ws_name, index, values  # 文件路径，工作表名，行数，行内容
            else:
                if space_line < 10:
                    space_line += 1
                else:
                    break


def find_excel_data(search_key: str, target_dir: str):
    """
    检索指定目录下的excel文件和过滤
    参数：
        search_key：检索的关键词
        target_dir：目标文件夹
    """
    for path, dirs, files in os.walk(target_dir):
        files = [file for file in files if not file.startswith('~$')]  # 过滤掉正打开的excel文件
        xls_files = [file for file in files if file.endswith('.xls')]  # 取出所有的xls后缀文件
        xlsx_files = [file for file in files if file.endswith('.xlsx')]  # 取出所有的xlsx后缀文件
        for xls in xls_files:
            for data in process_xls(path, xls):
                filepath, ws_name, index, values = data
                status = find_key(search_key, values)
                if status:
                    yield filepath, ws_name, index, values
        for xlsx in xlsx_files:
            for data in process_xlsx(path, xlsx):
                filepath, ws_name, index, values = data
                status = find_key(search_key, values)
                if status:
                    yield filepath, ws_name, index, values  # 输出内容：路径/文件名、工作表名、行数、行内容


if __name__ == '__main__':

    time1 = time.time()
    search_key = '刘家站垦殖场'
    target_dir = './'
    for data in find_excel_data(search_key, target_dir):
        print(list(data))
    time2 = time.time()
    print("\n程序运行结束，停止运行。耗时：{}秒".format(round(time2 - time1, 2)))
